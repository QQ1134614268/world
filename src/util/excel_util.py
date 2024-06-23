# -*- coding:utf-8 -*-
"""
@Time: 2021/12/26
@Description:
"""
import os
from datetime import datetime
from typing import Type

from openpyxl.reader.excel import load_workbook
from openpyxl.workbook import Workbook

from config.conf import DATE_FORMAT, RESOURCE_DIR
from config.enum_conf import ResponseCode
from config.env_default import world_env
from config.exception import WorldException
from config.log_conf import logger
from config.mysql_db import db


class Convert:
    # 读取excel类型 empty string number date boolean Error
    def __init__(self, nullable=False, max_length=20, comment="姓名", index=None, default=None):
        self.nullable = nullable
        self.max_length = max_length
        self.comment = comment
        self.index = index
        self.default = default

    @staticmethod
    def convert(value, **kwargs):
        return value


class DateConvert(Convert):

    @staticmethod
    def convert(value, **kwargs):
        if value is None:
            return None
        if isinstance(value, datetime):
            return value
        return datetime.strptime(value, DATE_FORMAT)


class FloatConvert(Convert):

    @staticmethod
    def convert(value, **kwargs):
        if value is None:
            if not kwargs.get("nullable"):
                raise WorldExcelException("转换类型异常")
            return kwargs.get("default")
        if isinstance(value, float):
            return value
        return float(value)


class StrConvert(Convert):
    # , nullable = True
    def convert(self, value, **kwargs):
        # nullable = False, max_length = 20, comment = "姓名", index = None
        if value is None:
            return None
        if isinstance(value, str):
            return value
        return str(value)


class Field:
    def __init__(self, convert_handle: Type[Convert], nullable=False, min_length=0, max_length=20, comment=None,
                 index=None, default=None):
        self.convert_handle = convert_handle
        self.comment = comment
        self.nullable = nullable
        self.min_length = min_length
        self.max_length = max_length
        self.index = index
        self.default = default

    def convert(self, value):
        return self.convert_handle.convert(value, nullable=self.nullable, max_length=self.max_length,
                                           default=self.default)


def check_excel_type(file_name):
    arr = file_name.split(".")
    file_type = arr and arr[-1]
    if file_type not in ["xlsx", "xls"]:
        raise WorldExcelException("上传文件格式不正确!")


class ExcelHandler:

    @staticmethod
    def get_sheet(fp=None):
        return load_workbook(fp).active

    @staticmethod
    def get_cn_en(cla_type=None):
        cn_en = {}
        for attr, field in cla_type.__dict__.items():
            if isinstance(field, Field):
                cn_en[field.comment] = attr
        return cn_en

    @staticmethod
    def get_cn(cla_type=None):
        cn_en = []
        for attr, field in cla_type.__dict__.items():
            if isinstance(field, Field):
                cn_en.insert(field.index, field.comment)
                # cn_en[field.index] = field.comment
        return cn_en

    @staticmethod
    def get_attr_map(cla_type=None):
        attr_map = {}
        for attr, field in cla_type.__dict__.items():
            if isinstance(field, Field):
                attr_map[attr] = field
        return attr_map

    @classmethod
    def get_excel_header(cls, rows):
        headers = []
        for row in rows:
            for cell in row:
                headers.append(cell.value)
            break
        return headers

    @classmethod
    def from_file(cls, fp=None, cla_type=None):
        errmsg = []
        ret = []
        sheet = cls.get_sheet(fp)
        rows = sheet.iter_rows()
        if sheet.max_row < 2:
            raise WorldExcelException("没有数据")
        excel_header = cls.get_excel_header(rows)
        obj_header = cls.get_cn(cla_type)
        if excel_header != obj_header:
            raise WorldExcelException(obj_header)
        attr_map = cls.get_attr_map(cla_type)
        for row_index, row in enumerate(rows):
            row_data = {}
            for col_index, attr_key in enumerate(attr_map.keys()):
                field = attr_map.get(attr_key)
                col_index = field.index or col_index
                cell = row[col_index - 1]
                try:
                    value = field.convert(cell.value)
                    row_data[attr_key] = value
                except WorldExcelException as e:
                    errmsg.append(f"{row_index + 2}行{col_index + 1}列, 值有问题")
                    logger.error(e)
            ret.append(row_data)
        if errmsg:
            raise WorldExcelException(message=errmsg)
        return ret

    @classmethod
    def to_file(cls, data=Type[list], cla_type=None, ):
        wb = Workbook()
        ws = wb.active
        attr_map = cls.get_attr_map(cla_type)
        for index, attr in enumerate(attr_map.items()):
            ws.cell(row=1, column=(index + 1)).value = attr[1].comment
        for row_index, item in enumerate(data):
            for col_index, attr in enumerate(attr_map.items()):
                if isinstance(item, dict):
                    ws.cell(row=(row_index + 2), column=(col_index + 1)).value = item.get(attr[0])
                if isinstance(item, db.Model):
                    ws.cell(row=(row_index + 2), column=(col_index + 1)).value = getattr(item, attr[0])
        return wb


class WorldExcelException(WorldException):

    def __init__(self, message, code=ResponseCode.EXCEL_FAIL.value):
        self.code = code
        self.message = message

    def __str__(self):
        return "code: %(code)d      message: %(message)s " % {'code': self.code, 'message': self.message}


if __name__ == '__main__':
    class User:
        name = Field(StrConvert, nullable=False, max_length=20, comment="姓名", index=1)
        sex = Field(StrConvert, nullable=False, max_length=20, comment="性别", index=2)
        birthday = Field(DateConvert, nullable=False, max_length=20, comment="生日", index=3)
    # 读取Excel
    path = os.path.join(RESOURCE_DIR, "excel_download_test.xlsx")
    dic_list = ExcelHandler.from_file(fp=path, cla_type=User)
    print(dic_list)
    # 写入Excel
    wb2 = ExcelHandler.to_file(data=dic_list, cla_type=User)
    wb2.save(os.path.join(world_env.log_dir, "test_excel_output.xlsx"))
