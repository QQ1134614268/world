# -*- coding:utf-8 -*-
"""
@Time: 2020/7/5
@Description: pass
"""

import datetime
import os
from enum import Enum

from flask import Blueprint, request
from openpyxl import load_workbook
from sqlalchemy import distinct

from config.conf import DATA_DIR
from config.mysql_db import db
from util import res_util
from vo.table_model import WorkerTime

work_api2 = Blueprint("work_api2", __name__, url_prefix='/api/work_api2')


class Kind(Enum):
    start = "上班"
    end = "下班"


@work_api2.route("/init_data", methods=['GET', 'POST'])
def init_data():
    name = ['7月打卡.xlsx', '8月打卡.xlsx', '9月打卡.xlsx', ]
    data = {}
    # session.query(User).filter(User.username == "Jack").delete()
    WorkerTime.query.delete()
    for file in name:
        wb = load_workbook(os.path.join(DATA_DIR, "work", file))
        table = wb['考勤统计汇总']
        morning_start = datetime.time(hour=7, minute=0, second=0)
        morning_end = datetime.time(hour=12, minute=0, second=0)
        afternoon_start = datetime.time(hour=13, minute=30, second=0)
        afternoon_end = datetime.time(hour=18, minute=0, second=0)
        evening_start = datetime.time(hour=19, minute=0, second=0)
        evening_end = datetime.time(hour=22, minute=30, second=0)
        day_end = datetime.time(hour=23, minute=29, second=59)
        for i in range(3, table.max_row):
            name = table.cell(row=i, column=2).value
            kind = table.cell(row=i, column=6).value
            note = table.cell(row=i, column=8).value
            att_datetime = table.cell(row=i, column=7).value
            att_datetime = datetime.datetime.strptime(att_datetime, '%Y-%m-%d %H:%M:%S')
            att_time = datetime.time(att_datetime.hour, att_datetime.minute, att_datetime.second, )

            key = (name, att_datetime.year, att_datetime.month, att_datetime.day)
            if key not in data:
                data[key] = WorkerTime(
                    name=name,
                    date=datetime.date(att_datetime.year, att_datetime.month, att_datetime.day),
                    note=note,
                    json_data=[],
                    json_data2=[]
                )

            vo = data[key]
            if att_time < morning_end and kind == Kind.start.value:
                if vo.morning_start:
                    vo.json_data.append({"time": att_time.strftime("%H:%M:%S"), "kind": kind})
                else:
                    vo.morning_start = att_time
            elif morning_start < att_time < afternoon_start and kind == Kind.end.value:
                if vo.morning_end:
                    vo.json_data.append({"time": att_time.strftime("%H:%M:%S"), "kind": kind})
                else:
                    vo.morning_end = att_time

            elif morning_end < att_time < afternoon_end and kind == Kind.start.value:
                if vo.afternoon_start:
                    vo.json_data.append({"time": att_time.strftime("%H:%M:%S"), "kind": kind})
                else:
                    vo.afternoon_start = att_time

            elif afternoon_start < att_time < evening_start and kind == Kind.end.value:
                if vo.afternoon_end:
                    vo.json_data.append({"time": att_time.strftime("%H:%M:%S"), "kind": kind})
                else:
                    vo.afternoon_end = att_time

            elif afternoon_end < att_time < evening_end and kind == Kind.start.value:
                if vo.evening_start:
                    vo.json_data.append({"time": att_time.strftime("%H:%M:%S"), "kind": kind})
                else:
                    vo.evening_start = att_time

            elif evening_start < att_time and kind == Kind.end.value:
                if vo.evening_end:
                    vo.json_data.append({"time": att_time.strftime("%H:%M:%S"), "kind": kind})
                else:
                    vo.evening_end = att_time
            else:
                vo.json_data2.append({"time": att_time.strftime("%H:%M:%S"), "kind": kind})
    db.session.add_all(data.values())  # 只能加一条数据
    db.session.commit()
    return res_util.success()


def _get_day_hours(vo):
    total = 0
    if vo.morning_start or vo.morning_end:
        total += 4.5
    if vo.afternoon_start or vo.afternoon_end:
        total += 4.5
    if vo.evening_start or vo.evening_end:
        total += 4.5
    return total


@work_api2.route("/get_worker", methods=['GET', 'POST'])
def get_worker():
    vos = db.session.query(distinct(WorkerTime.name)).all()
    ret = list(map(lambda x: x[0], vos))
    return res_util.success(ret)


@work_api2.route("/get_person", methods=['GET', 'POST'])
def get_person():
    name = request.args.get("name")
    year = request.args.get("year", type=int)
    month = request.args.get("month", type=int)
    day = request.args.get("day", type=int)

    # arr = [elem for elem in arr if elem != None]

    if day:
        day = datetime.date(year, month, day)
        vos = db.session.query(WorkerTime).filter(WorkerTime.name == name, WorkerTime.date == day).all()
        ret = []
        for vo in vos:
            ret.append({"date": vo.date.strftime("%Y-%m-%d"), "hours": _get_day_hours(vo)})
        return res_util.success(ret)
    if month:
        date = datetime.date(year, month, 1)
        last = datetime.date(year, month + 1, 1) - datetime.timedelta(days=1)
        vos = db.session.query(WorkerTime).filter(WorkerTime.date.between(date, last), WorkerTime.name == name).all()
    elif year:
        date = datetime.date(year, 1, 1)
        date.strftime("%Y-%m-%d")
        last = datetime.date(year + 1, 1, 1) - datetime.timedelta(days=1)
        vos = db.session.query(WorkerTime).filter(WorkerTime.date.between(date, last), WorkerTime.name == name).all()
    else:
        vos = db.session.query(WorkerTime).filter(WorkerTime.name == name).all()
    ret = _group_date(vos, year, month)
    ret.append({"date": "total", "hours": sum(map(lambda x: x["hours"], ret))})
    return res_util.success(ret)


def _group_date(vos, year, month):
    data = {}
    if month:
        for vo in vos:
            if vo.date.strftime("%Y-%m-%d") not in data:
                data[vo.date.strftime("%Y-%m-%d")] = []
            data[vo.date.strftime("%Y-%m-%d")].append(vo)
    elif year:
        for vo in vos:
            if vo.date.strftime("%Y-%m") not in data:
                data[vo.date.strftime("%Y-%m")] = []
            data[vo.date.strftime("%Y-%m")].append(vo)
    else:
        for vo in vos:
            if vo.date.strftime("%Y") not in data:
                data[vo.date.strftime("%Y")] = []
            data[vo.date.strftime("%Y")].append(vo)
    ret = []
    for key, value in data.items():
        ret.append({"date": key, "hours": sum(map(_get_day_hours, value))})
    return ret


@work_api2.route("/get_date", methods=['GET', 'POST'])
def get_date():
    year = request.args.get("year")
    month = request.args.get("month")
    day = request.args.get("day")
    if day:
        date = datetime.date(int(year), int(month), int(day))
        vos = db.session.query(WorkerTime).filter(WorkerTime.date == date).all()
    elif month:
        date = datetime.date(int(year), int(month), 1)
        last = datetime.date(int(year), int(month) + 1, 1) - datetime.timedelta(days=1)
        vos = db.session.query(WorkerTime).filter(WorkerTime.date.between(date, last)).all()
    else:
        date = datetime.date(int(year), 1, 1)
        last = datetime.date(int(year) + 1, 1, 1) - datetime.timedelta(days=1)
        vos = db.session.query(WorkerTime).filter(WorkerTime.date.between(date, last)).all()
    ret = _group_person(vos)
    ret.append({"name": "total", "hours": sum(map(lambda x: x["hours"], ret))})
    return res_util.success(ret)


def _group_person(vos):
    data = {}
    for vo in vos:
        if vo.name not in data:
            data[vo.name] = []
        data[vo.name].append(vo)
    ret = []
    for key, value in data.items():
        ret.append({"name": key, "hours": sum(map(_get_day_hours, value))})
    return ret


@work_api2.route("/get_all", methods=['GET', 'POST'])
def get_all():
    vos = db.session.query(WorkerTime).filter().all()
    ret = sum(map(_get_day_hours, vos))
    return res_util.success(ret)


@work_api2.route("/check", methods=['GET', 'POST'])
def check():
    # 上下班时间异常
    pass


@work_api2.route("/note_info", methods=['GET', 'POST'])
def note_info():
    # 备注信息--全部打卡信息
    pass
