# -*- coding:utf-8 -*-
"""
@Time: 2021/12/26
@Description:
"""
from datetime import date

from openpyxl import load_workbook

from config.exception import WorldException
from util import time_util


class Convert:
    # 读取excel类型
    # empty
    # string
    # number
    # date
    # boolean
    # Error
    def __init__(self, nullable=False, max_length=20, comment="姓名", index=None):
        self.nullable = nullable
        self.max_length = max_length
        self.comment = comment
        self.index = index

    @staticmethod
    def convert(value):
        return value


class DateConvert(Convert):

    @staticmethod
    def convert(src, nullable=None, max_length=None):
        if src is None:
            return None
        if isinstance(src, date):
            return src
        return time_util.get_datetime_by_str(src)


class FloatConvert(Convert):

    @staticmethod
    def convert(src, nullable):
        if src is None:
            return None
        if isinstance(src, float):
            return src
        return float(src)


class StrConvert(Convert):

    @staticmethod
    def convert(src, nullable):
        if src is None:
            return None
        if isinstance(src, str):
            return src
        return str(src)


# todo 重要
class Field:
    def __init__(self, obj_type: Convert, nullable=False, max_length=20, comment="姓名", index=None):
        self.obj_type = obj_type(nullable=nullable, max_length=max_length, comment=comment, index=index)
        self.comment = comment


class User:
    name = Field(StrConvert, nullable=False, max_length=20, comment="姓名", index=1)
    birthday = Field(DateConvert, nullable=False, max_length=20, comment="生日", index=1)
    pay = Field(FloatConvert, nullable=False, max_length=20, comment="薪资", index=1)


class ExcelExceptionMsg:
    pass


class ExcelHandler:
    # [姓名,生日]
    # 字段 field
    # convert
    # 赋值

    @staticmethod
    def get_sheet(fp=None):
        return load_workbook(fp).active

    @staticmethod
    def get_cn_en(cla_type=None):
        cn_en = {}
        for attr, field in cla_type.__dict__.items():
            if isinstance(field, Field):
                cn_en[field.comment] = attr
        return cn_en

    @staticmethod
    def get_attr_map(cla_type=None):
        attr_map = {}
        for attr, field in cla_type.__dict__.items():
            if isinstance(field, Field):
                attr_map[field.comment] = field
        return attr_map

    @classmethod
    def get_excel_header(cls, fp):
        headers = []
        sheet = cls.get_sheet(fp)
        for row in sheet.iter_rows():
            for cell in row:
                headers.append(cell.value)
            break
        return headers

    @classmethod
    def from_file(cls, fp=None, cla_type=None):
        errmsg = []
        ret = []
        sheet = cls.get_sheet(fp)
        excel_header = cls.get_excel_header(fp)
        cn_en = cls.get_cn_en(cla_type)
        attr_map = cls.get_attr_map(cla_type)
        for row_index, row in enumerate(sheet.iter_rows()):
            row_data = {}
            for col_index, cn_excel_attr in enumerate(excel_header):
                attr = cn_en.get(cn_excel_attr)
                field = attr_map.get(cn_excel_attr)
                cell = row[col_index]
                try:
                    value = field.obj_type.convert(cell.value)
                    row_data[attr] = value
                except:
                    errmsg.append(f"{row_index}行{col_index}列, 值有问题")
            ret.append(row_data)
        if errmsg:
            raise WorldException(errmsg)
        return ret

    def to_file(self):
        pass


if __name__ == '__main__':
    path = r'E:\world\resource\excel_download_test.xlsx'
    print(ExcelHandler.from_file(fp=path, cla_type=User))
