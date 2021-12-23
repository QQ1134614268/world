# -*- coding:utf-8 -*-
"""
@Time: 2020/7/5
@Description: pass
"""
from datetime import date, datetime
from io import BytesIO

import openpyxl
from flask import request, make_response, send_file, Blueprint
from flask_restful import Resource
from openpyxl import Workbook
from sqlalchemy import and_, func
from sqlalchemy.dialects.mysql import insert

from config.conf import header_dic
from config.mysql_db import db
from service.token_service import get_id_by_token
from util import res_util, db_util, time_util
from vo.table_model import WorkerVO, WorkerTimeVO

work_time_analyse_api = Blueprint("WorkTimeAnalyseApi", __name__, url_prefix='/api/work_api/WorkTimeAnalyseApi')


class WorkerExcelApi(Resource):
    """工人"""

    def post(self):
        file = request.files.get("file")
        wb = openpyxl.load_workbook(file)
        # sheet = wb["Sheet"]
        sheet = wb.active
        # get_column_letter column_index_from_string
        values = list(sheet.values)
        headers = values[0]
        header_dic_resvese = {v: k for k, v in header_dic.items()}
        en_header = [header_dic_resvese.get(header) for header in headers]
        data = []
        for v in values[1:]:
            data.append(dict(zip(en_header, v)))

        for val in data:
            val["belong"] = get_id_by_token()

        insert_stmt = insert(WorkerVO).values(data)
        on_duplicate_key_stmt = insert_stmt.on_duplicate_key_update(
            name=insert_stmt.inserted.name,
            birthday=insert_stmt.inserted.birthday,
            id_card_number=insert_stmt.inserted.id_card_number,
            start_time=insert_stmt.inserted.start_time,
            sex=insert_stmt.inserted.sex,
            phone=insert_stmt.inserted.phone,
            pay=insert_stmt.inserted.pay,
            belong=insert_stmt.inserted.belong,
        )
        try:
            db.session.execute(on_duplicate_key_stmt)
            db.session.commit()
            return res_util.success()
        except:
            return res_util.fail("请检查Excel内容")

    def get(self):
        vos = WorkerVO.query.filter(WorkerVO.belong == get_id_by_token()).all()
        wb = Workbook()
        ws = wb.active
        for index, header_name in enumerate(header_dic.values()):
            ws.cell(row=1, column=(index + 1)).value = header_name

        for row_index, vo in enumerate(vos):
            for col_index, attr in enumerate(header_dic.keys()):
                ws.cell(row=(row_index + 2), column=(col_index + 1)).value = getattr(vo, attr)
        byte_ios = BytesIO()
        wb.save(byte_ios)
        byte_ios.seek(0)
        filename = "名单.xlsx"
        response = make_response(
            send_file(byte_ios, as_attachment=True, attachment_filename=filename, mimetype='application/octet-stream'))
        response.headers.add("Cache control", "no cache")
        return response

    # obj.filename
    def check_excel_type(self,file_data):
        file_name = file_data.split(".")[0]
        file_type = file_data.split(".")[1]
        if not file_type in ["xlxs", "xls"]:
            raise
    # 读取excel类型
    # empty
    # string
    # number
    # date
    # boolean
    # Error
    def float_convert(self, src, nullable):
        if src is None:
            raise
        if isinstance(src, float):
            return src
        if isinstance(src, str):
            try:
                return float(src)
            except:
                raise

    def enum_convert(self, src, nullable=None, max_length=None):
        if src is None:
            raise
        if isinstance(src, float):
            return src
        if isinstance(src, str):
            try:
                return float(src)
            except:
                raise

    def date_convert(self, src, nullable=None, max_length=None):
        if src is None:
            raise
        if isinstance(src, date):
            return src
        if isinstance(src, str):
            try:
                return time_util.get_datetime_by_str(src)
            except:
                # 格式错误
                raise

    def date_convert(self, src, nullable=None, max_length=None):
        if src is None:
            raise
        if isinstance(src, datetime):
            return src
        if isinstance(src, str):
            try:
                return time_util.get_datetime_by_str(src)
            except:
                # 格式错误
                raise


# todo 重要
class Field:
    def __init__(self, obj_type, nullable=False, max_length=20, comment="姓名"):
        pass


class User:
    name = Field(str, nullable=False, max_length=20, comment="姓名")
    birthday = Field(datetime, nullable=False, max_length=20, comment="生日")
    pay = Field(float, nullable=False, max_length=20, comment="薪资")


class ExcelExceptionMsg:
    pass


class WorkerApi(Resource):
    """工人"""

    def post(self, _id):
        data = request.get_json()
        data["belong"] = get_id_by_token()
        vo = WorkerVO(**data)
        db.session.add(vo)
        db.session.commit()
        return res_util.success(vo.id)

    def get(self, _id):
        page = request.args.get("currentPage", 1, int)
        page_size = request.args.get("pageSize", 15, int)
        user_id = get_id_by_token()
        query_build = [WorkerVO.belong == user_id]
        if request.args.get("name"):
            query_build.append(WorkerVO.name.contains(request.args.get("name")))
        if request.args.get("phone"):
            query_build.append(WorkerVO.phone == request.args.get("phone"))
        if request.args.get("idCard"):
            query_build.append(WorkerVO.id_card_number == request.args.get("idCard"))
        if request.args.get("startDate"):
            query_build.append(WorkerVO.start_time >= request.args.get("startDate"))
        if request.args.get("endDate"):
            query_build.append(WorkerVO.start_time <= request.args.get("endDate"))
        vos = WorkerVO.query.filter(*query_build).paginate(page=page, per_page=page_size)
        return res_util.page_success(vos)

    def put(self, _id):
        data = request.get_json()
        WorkerVO.query.filter(WorkerVO.id == _id).update(data)
        db.session.commit()
        return res_util.success(_id)

    def delete(self, _id):
        WorkerVO.query.filter(WorkerVO.id == _id).delete()
        db.session.commit()
        return res_util.success()


class WorkerTimeApi(Resource):
    """工时"""

    def post(self, _id):
        data = request.get_json()
        db.session.add(WorkerTimeVO(**data))
        db.session.commit()
        return res_util.success()

    def get(self, _id):
        date = request.args.get("date")
        name = request.args.get("name")
        page = request.args.get("currentPage", 1, int)
        page_size = request.args.get("pageSize", 15, int)
        user_id = get_id_by_token()

        query_filter = [WorkerVO.belong == user_id]
        if name:
            query_filter.append(WorkerVO.name.contains(name))

        join_filter = [WorkerTimeVO.worker_id == WorkerVO.id]
        if date:
            join_filter.append(WorkerTimeVO.date == date)
        res = WorkerVO.query.outerjoin(
            WorkerTimeVO,
            and_(*join_filter)
        ).with_entities(
            WorkerVO.id,
            WorkerVO.name,
            WorkerTimeVO.date,
            WorkerTimeVO.morning,
            WorkerTimeVO.noon,
            WorkerTimeVO.afternoon,
            WorkerTimeVO.night,
        ).filter(*query_filter).paginate(page=page, per_page=page_size)
        return res_util.page_success_row(res)

    def put(self, _id):
        data = request.get_json()
        update_data = {
            'date': data.get("date"),
            'worker_id': data.get("worker_id"),
        }
        if data.get("type") == 'morning':
            update_data['morning'] = data.get("flag")
            update_data['morning_area'] = data.get("area")
            update_data['morning_content'] = data.get("content")
        if data.get("type") == 'noon':
            update_data['noon'] = data.get("flag")
            update_data['noon_area'] = data.get("area")
            update_data['noon_content'] = data.get("content")
        if data.get("type") == 'afternoon':
            update_data['afternoon'] = data.get("flag")
            update_data['afternoon_area'] = data.get("area")
            update_data['afternoon_content'] = data.get("content")
        if data.get("type") == 'night':
            update_data['night'] = data.get("flag")
            update_data['night_area'] = data.get("area")
            update_data['night_content'] = data.get("content")
        u_list = [(k, v) for k, v in update_data.items()]
        insert_stmt = insert(WorkerTimeVO).values(update_data)
        on_duplicate_key_stmt = insert_stmt.on_duplicate_key_update(u_list)
        db.session.execute(on_duplicate_key_stmt)
        db.session.commit()
        return res_util.success()


class WorkTimeAnalyseApi(Resource):
    """工时统计"""

    @staticmethod
    @work_time_analyse_api.route('/get_sum_time/<int:_id>', methods=['GET'])
    def get_sum_time(_id):
        query_filter = [WorkerVO.belong == get_id_by_token()]
        date = request.args.getlist("date[]")
        if date:
            query_filter.extend([WorkerTimeVO.date >= date[0], WorkerTimeVO.date < date[1]])
        name = request.args.get("name")
        if name:
            query_filter.append(WorkerVO.name.contains(name))

        res = WorkerVO.query.outerjoin(
            WorkerTimeVO,
            and_(WorkerTimeVO.worker_id == WorkerVO.id)
        ).with_entities(
            WorkerVO.id,
            WorkerVO.name,
            WorkerVO.pay,
            func.sum(WorkerTimeVO.morning).label('morning'),
            func.sum(WorkerTimeVO.noon).label('noon'),
            func.sum(WorkerTimeVO.afternoon).label('afternoon'),
            func.sum(WorkerTimeVO.night).label('night'),
        ).filter(
            *query_filter
        ).group_by(
            WorkerVO.id
        ).all()
        data = db_util.row_to_dic(res)
        for item in data:
            item["days"] = (item.get("morning") * 4.5 + item.get("noon") * 2 + item.get("afternoon") * 4.5 + item.get(
                "night") * 4.5) / 9
            item["money"] = item["days"] * item["pay"]
        return res_util.json_success(data)

    @staticmethod
    @work_time_analyse_api.route('/get_day_report/<int:_id>', methods=['GET'])
    def get_day_report(_id):
        # 日报模式

        return res_util.json_success()
