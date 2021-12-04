# -*- coding:utf-8 -*-
"""
@Time: 2020/7/5
@Description: pass
"""
import time
from io import BytesIO

import openpyxl
from flask import request, make_response, send_file
from flask_restful import Resource
from openpyxl import Workbook
from sqlalchemy.dialects.mysql import insert

import service
from config.mysql_db import db
from service import worker_service
from util import res_util
from vo.table_model import WorkerVO

header_dic = {
    "name": "姓名",
    "id_card_number": "身份证",
    "phone": "电话",
    "sex": "性别",
    "birthday": "生日",
    "start_time": "入职日期",
    "pay": "薪资",
}


class WorkerExcelApi(Resource):
    """工人"""

    def post(self):
        file = request.files.get("file")
        wb = openpyxl.load_workbook(file)
        # sheet = wb["Sheet"]
        sheet = wb.active
        # get_column_letter column_index_from_string
        values = list(sheet.values)
        headers = values[0]
        header_dic_resvese = {v: k for k, v in header_dic.items()}
        en_header = [header_dic_resvese.get(header) for header in headers]
        data = []
        for v in values[1:]:
            data.append(dict(zip(en_header, v)))

        for val in data:
            val["belong"] = service.token_service.get_id_by_token()
        insert_stmt = insert(WorkerVO).values(data)
        on_duplicate_key_stmt = insert_stmt.on_duplicate_key_update(
            name=insert_stmt.inserted.name,
            birthday=insert_stmt.inserted.birthday,
            id_card_number=insert_stmt.inserted.id_card_number,
            start_time=insert_stmt.inserted.start_time,
            sex=insert_stmt.inserted.sex,
            phone=insert_stmt.inserted.phone,
            pay=insert_stmt.inserted.pay,
            belong=insert_stmt.inserted.belong,
        )
        db.session.execute(on_duplicate_key_stmt)
        db.session.commit()
        return res_util.success()

    def get(self):
        vos = WorkerVO.query.filter(WorkerVO.belong == service.token_service.get_id_by_token()).all()
        wb = Workbook()
        ws = wb.active
        for index, header_name in enumerate(header_dic.values()):
            ws.cell(row=1, column=(index + 1)).value = header_name

        for row_index, vo in enumerate(vos):
            for col_index, attr in enumerate(header_dic.keys()):
                ws.cell(row=(row_index + 2), column=(col_index + 1)).value = getattr(vo, attr)
        byte_ios = BytesIO()
        wb.save(byte_ios)
        byte_ios.seek(0)
        filename = "名单.xlsx"
        response = make_response(
            send_file(byte_ios, as_attachment=True, attachment_filename=filename, mimetype='application/octet-stream'))
        response.headers.add("Cache control", "no cache")
        return response


class WorkerApi(Resource):
    """工人"""

    def post(self, _id):
        data = request.get_json()
        data["belong"] = service.token_service.get_id_by_token()
        vo = WorkerVO(**data)
        db.session.commit()
        return res_util.success(vo.id)

    def get(self, _id):
        page = request.args.get("currentPage", 1, int)
        page_size = request.args.get("pageSize", 15, int)
        user_id = service.token_service.get_id_by_token()
        query_build = [WorkerVO.belong == user_id]
        if request.args.get("name"):
            query_build.append(WorkerVO.name.contains(request.args.get("name")))
        if request.args.get("phone"):
            query_build.append(WorkerVO.phone == request.args.get("phone"))
        if request.args.get("idCard"):
            query_build.append(WorkerVO.id_card_number == request.args.get("idCard"))
        if request.args.get("startDate"):
            query_build.append(WorkerVO.start_time >= request.args.get("startDate"))
        if request.args.get("endDate"):
            query_build.append(WorkerVO.start_time <= request.args.get("endDate"))
        vos = WorkerVO.query.filter(*query_build).paginate(page=page, per_page=page_size)
        return res_util.page_success(vos)

    def put(self, _id):
        data = request.get_json()
        WorkerVO.query.filter(id=_id).update(**data)
        db.session.commit()
        return res_util.success(_id)

    def delete(self, _id):
        WorkerVO.query.filter(WorkerVO.id == _id).delete()
        db.session.commit()
        return res_util.success()


class WorkerTimeApi(Resource):
    """工时"""

    def post(self):
        data = request.get_json()

        if isinstance(data.get("worker_id"), list):
            return worker_service.cover_worker_time3(data)
        if isinstance(data.get("type"), list):
            return worker_service.cover_worker_time2(data)

        worker_service.cover_worker_time(data)
        return worker_service.add_worker_time(data)

    def get(self, _id):
        if request.args.get("month"):
            return res_util.success([])
            # return worker_service.get_worker_month(request.args.get("month"), request.args.get("workerId"))
        date = request.args.get("date", time.strftime('%Y-%m-%d 00:00:00'))
        return worker_service.get_worker_day(date)

    def put(self):
        data = request.get_json()
        return worker_service.update_worker_time(data)
